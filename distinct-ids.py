import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import logging

def sanitizeForExcel(value):
    if isinstance(value, (set, list, tuple)):
        return ', '.join(str(item) for item in value)
    return str(value) if value is not None else ''

logging.basicConfig(level=logging.DEBUG)
projectsDict = defaultdict(lambda: defaultdict(set))
updatedRows = set()
logging.info("Started")

try:
    wbInput = load_workbook(filename='output/filled.xlsx', read_only=False)
    logging.info("Loaded")
    sheetInput = wbInput.active
    for rowIndex, row in enumerate(sheetInput.iter_rows(min_row=2, values_only=True), start=2):
        idValue = str(row[20].strip()) if row[20] is not None else ""
        existingId = projectsDict.get(idValue, None)
        if existingId is None:
            projectsDict[idValue] = rowIndex
            updatedRows.add(int(rowIndex))
except Exception as e:
    logging.error(e)
    logging.error("Error loading input")

logging.info("Finished Processed")

wbUpdated = Workbook()
sheetUpdated = wbUpdated.active
header = [cell.value for cell in sheetInput[1]]
sheetUpdated.append(header)

for row in sheetInput.iter_rows(min_row=2, values_only=False):
    rowNumber = row[0].row
    if rowNumber in updatedRows:
        updatedRow = [cell.value for cell in row]
        sanitizedRow = [sanitizeForExcel(cell) for cell in updatedRow]
        sheetUpdated.append(sanitizedRow)

wbUpdated.save('output/distinct.xlsx')
logging.info("Results have been saved to output/distinct.xlsx")