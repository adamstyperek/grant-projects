import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import Levenshtein
import logging

logging.basicConfig(level=logging.DEBUG)

projectNamesItems = defaultdict(lambda: defaultdict(set))
acronymsItems = defaultdict(lambda: defaultdict(set))

try:
    wbTitles = load_workbook(filename='data/slownik_tytulow DBI.xlsx', read_only=False)
    sheetTitlesInput = wbTitles.active
    for rowIndex, row in enumerate(sheetTitlesInput.iter_rows(min_row=2, values_only=True), start=2):
        normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', str(row[21]).strip().lower()) if row[21] is not None else ""
        normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', str(row[22]).strip().lower()) if row[22] is not None else ""
        projectName = str(row[-1]).strip() if row[-1] is not None else ""
        reportNumber = str(row[0]).strip() if row[-1] is not None else ""
        index = normalisedProjectName + ";" + normalisedAcronym + ";" + reportNumber

        if projectName:
            projectNamesItems[index]['acronymName'] = normalisedAcronym
            projectNamesItems[index]['projectName'] = normalisedProjectName
            projectNamesItems[index]['projectNameToUse'] = projectName
            projectNamesItems[index]['reportNumber'] = reportNumber
except Exception as e:
    logging.error(f"An error occurred while processing row: {e}")

try:
    wbAcronyms = load_workbook(filename='data/slownik_akronimow DBI.xlsx', read_only=False)
    sheetAcronymsInput = wbAcronyms.active
    for rowIndex, row in enumerate(sheetAcronymsInput.iter_rows(min_row=2, values_only=True), start=2):
        normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', str(row[21]).strip().lower()) if row[21] is not None else ""
        normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', str(row[22]).strip().lower()) if row[22] is not None else ""
        acronym = str(row[-1]).strip() if row[-1] is not None else ""
        reportNumber = str(row[0]).strip() if row[-1] is not None else ""
        index = normalisedProjectName + ";" + normalisedAcronym + ";" + reportNumber

        if projectName:
            acronymsItems[index]['acronymName'] = normalisedAcronym
            acronymsItems[index]['projectName'] = normalisedProjectName
            acronymsItems[index]['acronym'] = acronym
            acronymsItems[index]['reportNumber'] = reportNumber
except Exception as e:
    logging.error(f"An error occurred while processing row: {e}")

# Create output workbook
wbProjectNamesOutput = Workbook()
sheetProjectNamesOutput = wbProjectNamesOutput.active
sheetProjectNamesOutput.title = "Projects"
sheetProjectNamesOutput.append(["Project Name", "Acronym", "Project name to use", "Report number"])
# Write header
for index, data in projectNamesItems.items():
    reportNumber = data.get('reportNumber', '')
    acronym = data.get('acronymName', '')
    projectName = data.get('projectName', '')
    projectNameToUse = data.get('projectNameToUse', '')
    sheetProjectNamesOutput.append([projectName, acronym, projectNameToUse, reportNumber])

wbProjectNamesOutput.save('data/dictionaries/project_names/manual.xlsx')

print(f"Results have been saved to data/dictionaries/project_names/manual.xlsx")

# Create output workbook
wbAcronymsOutput = Workbook()
sheetAcronymsOutput = wbAcronymsOutput.active
sheetAcronymsOutput.title = "acronyms"
sheetAcronymsOutput.append(["Project Name", "Acronym", "Acronym to use", "Report number"])
# Write header
for index, data in acronymsItems.items():
    reportNumber = data.get('reportNumber', '')
    acronym = data.get('acronymName', '')
    projectName = data.get('projectName', '')
    acronymToUse = data.get('acronym', '')
    sheetAcronymsOutput.append([projectName, acronym, acronymToUse, reportNumber])

wbAcronymsOutput.save('data/dictionaries/acronyms/manual.xlsx')

print(f"Results have been saved to data/dictionaries/acronyms/manual.xlsx")