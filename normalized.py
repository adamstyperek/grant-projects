import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict

# Load the workbook
wb = load_workbook(filename='data/input_e.xlsx')
sheet = wb.active

# Get the current number of columns
max_column = sheet.max_column

# Add two new column headers
sheet.cell(row=1, column=max_column + 1, value="Normalized Project Name")
sheet.cell(row=1, column=max_column + 2, value="Normalized Acronym")

# Process the data
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    project_name = re.sub(r'[^a-zA-Z0-9]', '', str(row[21].value).strip().lower()) if row[21].value else ""
    acronym = re.sub(r'[^a-zA-Z0-9]', '', str(row[22].value).strip().lower()) if row[22].value else ""

    # Write normalized values to new columns
    row[max_column].value = project_name
    row[max_column + 1].value = acronym

# Save the modified workbook
wb.save('data/input_e_modified.xlsx')

print("Modified file has been saved as data/input_e_modified.xlsx")