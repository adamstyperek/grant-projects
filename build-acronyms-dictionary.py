from openpyxl import load_workbook, Workbook
from collections import defaultdict
from Levenshtein import distance
import logging

logging.basicConfig(level=logging.DEBUG)

projects = defaultdict(lambda: defaultdict(set))

# Read input data
wbInput = load_workbook(filename='data/input.xlsx', read_only=False)
sheetInput = wbInput.active

for rowIndex, row in enumerate(sheetInput.iter_rows(min_row=2, values_only=True), start=2):
    projectName = str(row[21]).strip() if row[21] is not None else ""
    normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', projectName.lower())
    acronym = str(row[22]).strip() if row[22] is not None else ""
    normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', acronym.lower())
    rowNumber = str(rowIndex)
    projectNameToNormalized[projectName] = normalisedProjectName
    projectNumber = str(row[0]).strip() if row[0] is not None else ""

    if projectName and acronym:
        projects[normalisedProjectName]['projectName'] = projectName
        projects[normalisedProjectName]['acronyms'].add(acronym)
        projects[normalisedProjectName]['normalisedAcronyms'].add(normalisedAcronym)
        projects[normalisedProjectName]['rowNumbers'].add(rowNumber)
        projects[normalisedProjectName]['projectNumbers'] = projectNumber

sheetInput.insert_cols(sheetInput.max_column + 1, 4)
sheetInput.cell(row=1, column=sheetInput.max_column - 3, value="Nazwa projektu")
sheetInput.cell(row=1, column=sheetInput.max_column - 2, value="Sugerowany akronim")
sheetInput.cell(row=1, column=sheetInput.max_column - 1, value="Aktualny akronim")
sheetInput.cell(row=1, column=sheetInput.max_column, value="Wybrany akronim")

# Create output workbook
wbOutput = Workbook()
sheetOutput = wbOutput.active
sheetOutput.title = "Projects"

# Create a set to store the row numbers of updated rows
updatedRows = set()

# Write header
sheetOutput.append(["Source ID", "Project Name", "First Acronym", "Acronym", "Row Number", "Distance from First Acronym"])

# Write data and update input sheet
for project, data in projects.items():
    acronyms = list(data['acronyms'])
    normalisedAcronyms = list(data['normalisedAcronyms'])
    rowNumbers = list(data['rowNumbers'])
    projectName = data.get('projectName', '')
    projectNumbers = list(data['projectNumbers'])

    if len(acronyms) > 1:
        firstAcronym = acronyms[0]
        firstNormalisedAcronym = normalisedAcronyms[0]
        for i, (acronym, normalisedAcronym, rowNumber, projectNumber) in enumerate(zip(acronyms, normalisedAcronyms, rowNumbers, projectNumbers)):
            dist = distance(firstNormalisedAcronym, normalisedAcronym)
            sourceId = projectNumber
            sheetOutput.append([sourceId, projectName, firstAcronym, acronym, rowNumber, dist])
            if i > 0 and dist >= 1:
                # Update input sheet
                inputRow = sheetInput[rowNumber]
                inputRow[-4].value = projectName
                inputRow[-3].value = firstAcronym
                inputRow[-2].value = acronym
                inputRow[-1].value = ""
                updatedRows.add(int(rowNumber))

# Create a new workbook for the updated input data
wbUpdated = Workbook()
sheetUpdated = wbUpdated.active

# Copy header from the original input sheet
header = [cell.value for cell in sheetInput[1]]
sheetUpdated.append(header)

# Copy only the updated rows to the new sheet, with corrected data
for row in sheetInput.iter_rows(min_row=2, values_only=False):
    rowNumber = row[0].row
    if rowNumber in updatedRows:
        updatedRow = [cell.value for cell in row]
        normalizedName = projectNameToNormalized.get(str(updatedRow[-4]).strip(), str(updatedRow[-4]).strip())

        logging.debug(f"Row {rowNumber}: normalizedName = {normalizedName}")
        logging.debug(f"Projects dict for this name: {projects[normalizedName]}")

        logging.debug(f"Updated row: {updatedRow}")

        sanitizedRow = [sanitize_for_excel(cell) for cell in updatedRow]
        sheetUpdated.append(sanitizedRow)

# Save the workbooks
wbOutput.save('data/projects3.xlsx')
wbUpdated.save('data/input_e_modified_updated1.xlsx')

print("Results have been saved to data/projects.xlsx")
print("Updated input data (only modified rows) saved to data/input_e_modified_updated.xlsx")
print("Updated input data saved to data/input_e_modified_updated.xlsx")