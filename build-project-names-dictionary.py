import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import Levenshtein
import logging

logging.basicConfig(level=logging.DEBUG)

acronyms = defaultdict(lambda: defaultdict(set))

wb = load_workbook(filename='data/input.xlsx', read_only=False)
sheetInput = wb.active
for rowIndex, row in enumerate(sheetInput.iter_rows(min_row=2, values_only=True), start=2):
    normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', str(row[21]).strip().lower()) if row[21] is not None else ""
    normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', str(row[22]).strip().lower()) if row[22] is not None else ""
    projectName = str(row[21]).strip() if row[21] is not None else ""
    acronym = str(row[22]).strip() if row[22] is not None else ""

    if projectName and acronym:
        acronyms[normalisedAcronym]['acronymName'] = acronym
        acronyms[normalisedAcronym]['projects'].add(projectName)
        acronyms[normalisedAcronym]['normalisedProjects'].add(normalisedProjectName)

# Create output workbook
wbOutput = Workbook()
sheetOutput = wbOutput.active
sheetOutput.title = "acronyms"
updatedRows = set()
sheetOutput.append(["Project Name", "Acronym", "Project name to use"])
# Write header
for acronym, data in acronyms.items():
    projects = list(data['projects'])
    normalisedProjects = list(data['normalisedProjects'])
    projectsCount = len(projects)
    acronymName = data.get('acronymName', '')

    if projectsCount > 1:
        firstProject = projects[0]
        firstNormalisedProject = normalisedProjects[0]
        for i, (project, normalisedProjectName) in enumerate(zip(projects, normalisedProjects)):
            dist = Levenshtein.distance(firstProject, project)
            if i > 0 and dist < 50:
                sheetOutput.append([normalisedProjectName, acronym, firstProject])

wbOutput.save('data/dictionaries/project_names/based-on-distance.xlsx')

print(f"Results have been saved to data/dictionaries/project_names/based-on-distance.xlsx")