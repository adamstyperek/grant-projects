import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import logging

def findAcronymInManualDictionaries(projectIndex, acronym):
    returnedValue = acronym
    for index, acronymToUse in manual_acronyms_dictionary.items():
        if index == projectIndex:
            returnedValue = acronymToUse
    return returnedValue

def findProjectNameInManualDictionaries(projectIndex, projectName):
    returnedValue = projectName
    for index, projectNameToUse in project_names_dictionary.items():
        if index == projectIndex:
            returnedValue = projectNameToUse
            break
    return returnedValue

def findProjectNameInDictionaries(projectIndex, projectName):
    returnedValue = projectName
    for index, projectNameToUse in project_names_dictionary.items():
        if index == projectIndex:
            returnedValue = projectNameToUse
            break
    return returnedValue

logging.basicConfig(level=logging.DEBUG)
projects = defaultdict(lambda: defaultdict(set))
indexes = defaultdict(lambda: defaultdict(set))
project_names_dictionary = defaultdict(lambda: defaultdict(set))
manual_acronyms_dictionary = defaultdict(lambda: defaultdict(set))
manual_project_names_dictionary = defaultdict(lambda: defaultdict(set))
projects = defaultdict(lambda: defaultdict(set))




try:
    wb_dictionary_distance_project_names = load_workbook(filename='data/dictionaries/project_names/based-on-distance.xlsx', read_only=False)
    dictionary_distance_project_names_input = wb_dictionary_distance_project_names.active
    for row_index, row in enumerate(dictionary_distance_project_names_input.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0]) + ';' + str(row[1])
        project_names_dictionary[index] = str(row[2])
except:
    logging.error("Error loading dictionary_distance_project_names_input")


try:
    wb_dictionary_manual_project_names = load_workbook(filename='data/dictionaries/project_names/manual.xlsx', read_only=False)
    dictionary_manual_project_names_input = wb_dictionary_manual_project_names.active
    for row_index, row in enumerate(dictionary_manual_project_names_input.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0]) + ';' + str(row[1]) + ';' + str(row[2])
        manual_project_names_dictionary[index] = str(row[3])
except:
    logging.error("Error loading dictionary_manual_project_names_input")

try:
    wb_dictionary_manual_acronyms = load_workbook(filename='data/dictionaries/acronyms/manual.xlsx', read_only=False)
    dictionary_manual_acronyms_input = wb_dictionary_manual_acronyms.active
    for row_index, row in enumerate(dictionary_manual_acronyms_input.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0]) + ';' + str(row[1]) + ';' + str(row[2])
        manual_acronyms_dictionary[index] = str(row[3])
except:
    logging.error("Error loading dictionary_manual_acronyms_input")

try:
    print(project_names_dictionary)
    print("Project names dict length: " + str(len(project_names_dictionary)))
    print("Manual Project names dict length: " + str(len(manual_project_names_dictionary)))
    print("Manual Acronyms dict length: " + str(len(manual_acronyms_dictionary)))
    wb_input = load_workbook(filename='data/input.xlsx', read_only=False)
    print("Loaded")
    sheet_input = wb_input.active
    for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, values_only=True), start=2):
        report_number = str(row[0]).strip() if row[0] is not None else ""
        normalised_project_name = re.sub(r'[^a-zA-Z0-9]', '', str(row[21]).strip().lower()) if row[21] is not None else ""
        normalised_acronym = re.sub(r'[^a-zA-Z0-9]', '', str(row[22]).strip().lower()) if row[22] is not None else ""
        project_index_for_manual = normalised_project_name + ';' + normalised_acronym + ';' + report_number
        project_name = str(row[21]).strip() if row[21] is not None else ""
        acronym = str(row[22]).strip() if row[22] is not None else ""
        project_name = findProjectNameInManualDictionaries(project_index_for_manual, project_name)
        acronym = findAcronymInManualDictionaries(project_index_for_manual, acronym)
        normalised_project_name = re.sub(r'[^a-zA-Z0-9]', '', str(project_name).strip().lower())
        normalised_acronym = re.sub(r'[^a-zA-Z0-9]', '', str(acronym).strip().lower())
        project_index = normalised_project_name + ';' + normalised_acronym
        project_name = findProjectNameInDictionaries(project_index, project_name)
        normalised_project_name = re.sub(r'[^a-zA-Z0-9]', '', str(project_name).strip().lower())
        existing_index = indexes.get(project_index, None)
        newIndex = len(indexes) + 1
        if existing_index is None:
            indexes[project_index]['index'] = newIndex
            indexes[project_index]['project_name'] = project_name
            indexes[project_index]['acronym'] = acronym
except Exception as e:
    logging.error(e)
    logging.error("Error loading input")

wb_output = Workbook()
sheet_output = wb_output.active
sheet_output.title = "Indexes"

sheet_output.append(["Project Index", "Project Name", "Acronym", "Index"])
print(indexes)
print("Indexes length: " + str(len(indexes)))
for iterator, data in indexes.items():
    projectIndex = data.get('index', '')
    projectName = data.get('project_name', '')
    acronym = data.get('acronym', '')
    sheet_output.append([iterator, projectName, acronym, projectIndex])

wb_output.save('data/dictionaries/index/index.xlsx')
print("Results have been saved to data/dictionaries/index/index.xlsx")