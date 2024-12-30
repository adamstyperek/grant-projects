from openpyxl import load_workbook, Workbook
from collections import defaultdict
from Levenshtein import distance
import logging

logging.basicConfig(level=logging.DEBUG)

projects = defaultdict(lambda: defaultdict(set))
project_name_to_normalized = {}
def sanitize_for_excel(value):
    if isinstance(value, (set, list, tuple)):
        return ', '.join(str(item) for item in value)
    return str(value) if value is not None else ''

# Read input data
wb_input = load_workbook(filename='data/input_e_modified.xlsx', read_only=False)
sheet_input = wb_input.active

for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, values_only=True), start=2):
    project_name = str(row[21]).strip() if row[21] is not None else ""
    normalised_project_name = str(row[-2]).strip() if row[-2] is not None else ""
    acronym = str(row[22]).strip() if row[22] is not None else ""
    normalised_acronym = str(row[-1]).strip() if row[-1] is not None else ""
    institute = str(row[18]).strip() if row[18] is not None else ""
    project_index = normalised_project_name + ';' + normalised_acronym + ";" + institute
    row_number = str(row_index)
    project_name_to_normalized[project_name] = normalised_project_name
    project_number = str(row[0]).strip() if row[0] is not None else ""
    status = str(row[24]).strip() if row[24] is not None else ""


    if project_index and status.lower() in ["zakończony", "odrzucony formalnie", "odrzucony merytorycznie", "wycofany", "przerwany", "porzucony"]:
        if project_index not in projects:
            projects[project_index] = defaultdict(list)
        projects[project_index]['project_names'].append(project_name)
        projects[project_index]['acronyms'].append(acronym)
        projects[project_index]['statuses'].append(status)
        projects[project_index]['institutes'].append(institute)
        projects[project_index]['normalised_acronyms'].append(normalised_acronym)
        projects[project_index]['row_numbers'].append(row_number)
        projects[project_index]['project_numbers'].append(project_number)
sheet_input.insert_cols(sheet_input.max_column + 1, 6)
sheet_input.cell(row=1, column=sheet_input.max_column - 5, value="Nazwa projektu")
sheet_input.cell(row=1, column=sheet_input.max_column - 4, value="Sugerowany akronim")
sheet_input.cell(row=1, column=sheet_input.max_column - 3, value="Aktualny akronim")
sheet_input.cell(row=1, column=sheet_input.max_column - 2, value="Wybrany akronim")
sheet_input.cell(row=1, column=sheet_input.max_column - 1, value="Status")
sheet_input.cell(row=1, column=sheet_input.max_column, value="Instytut")
# Create output workbook
wb_output = Workbook()
sheet_output = wb_output.active
sheet_output.title = "Projects"
# Write header
# Create a set to store the row numbers of updated rows
updated_rows = set()

# Write header
sheet_output.append(["Source ID", "Project Name", "First Acronym", "Acronym", "Row Number", "Distance from First Acronym", "Status"])
more_than_one_acronyms = 0
one_acronym = 0
with_status = 0
# Write data and update input sheet
for project, data in projects.items():
    acronyms = (data['acronyms'])
    normalised_acronyms = (data['normalised_acronyms'])
    row_numbers = (data['row_numbers'])
    project_names = (data['project_names'])
    project_numbers = (data['project_numbers'])
    statuses = (data['statuses'])
    institutes = (data['institutes'])

    if len(acronyms) > 1:
        more_than_one_acronyms += 1
        first_acronym = acronyms[0]
        first_normalised_acronym = normalised_acronyms[0]
        for i, (acronym, normalised_acronym, row_number, project_number, status, institute, project_name) in enumerate(zip(acronyms, normalised_acronyms, row_numbers, project_numbers, statuses, institutes, project_names)):
            dist = distance(first_normalised_acronym, normalised_acronym)
            source_id = project_number
            sheet_output.append([source_id, project_name, first_acronym, acronym, row_number, dist, status])
            if status.lower() in ["zakończony", "odrzucony formalnie", "odrzucony merytorycznie", "wycofany", "przerwany", "porzucony"]:
                with_status += 1
                # Update input sheet
                input_row = sheet_input[row_number]
                input_row[-6].value = project_name
                input_row[-5].value = first_acronym
                input_row[-4].value = acronym
                input_row[-3].value = ""
                input_row[-2].value = status
                input_row[-1].value = institute
                updated_rows.add(int(row_number))
    else:
        one_acronym += 1

# Create a new workbook for the updated input data
wb_updated = Workbook()
sheet_updated = wb_updated.active

# Copy header from the original input sheet
header = [cell.value for cell in sheet_input[1]]
sheet_updated.append(header)

# Copy only the updated rows to the new sheet, with corrected data
for row in sheet_input.iter_rows(min_row=2, values_only=False):
    row_number = row[0].row
    if row_number in updated_rows:
        updated_row = [cell.value for cell in row]
        normalized_name = project_name_to_normalized.get(str(updated_row[-4]).strip(), str(updated_row[-4]).strip())
        sanitized_row = [sanitize_for_excel(cell) for cell in updated_row]
        sheet_updated.append(sanitized_row)
# Save the workbooks
wb_output.save('data/projects_with_statuses.xlsx')
wb_updated.save('data/input_e_modified_updated_with_statuses.xlsx')

print("Projects length:", len(projects))
print("More than one acronyms:", more_than_one_acronyms)
print("One acronyms:", one_acronym)
print("With status:", with_status)
print("Results have been saved to data/projects.xlsx")
print("Updated input data (only modified rows) saved to data/input_e_modified_updated.xlsx")
print("Updated input data saved to data/input_e_modified_updated.xlsx")