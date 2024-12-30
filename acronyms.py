from openpyxl import load_workbook, Workbook
from collections import defaultdict
import Levenshtein
import logging

logging.basicConfig(level=logging.DEBUG)

acronyms = defaultdict(lambda: defaultdict(set))
acronymToNormalized = {}

def sanitizeForExcel(value):
    if isinstance(value, (set, list, tuple)):
        return ', '.join(str(item) for item in value)
    return str(value) if value is not None else ''

wb = load_workbook(filename='data/input.xlsx', read_only=False)
sheetInput = wb.active
for rowIndex, row in enumerate(sheetInput.iter_rows(min_row=2, values_only=True), start=2):
    projectName = str(row[21]).strip() if row[21] is not None else ""
    normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', projectName.lower())
    acronym = str(row[22]).strip() if row[22] is not None else ""
    normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', acronym.lower())
    rowNumber = str(rowIndex)
    acronymToNormalized[acronym] = normalisedAcronym
    projectNumber = str(row[0]).strip() if row[0] is not None else ""

    if projectName and acronym:
        acronyms[normalisedAcronym]['acronymName'] = acronym
        acronyms[normalisedAcronym]['projects'].add(projectName)
        acronyms[normalisedAcronym]['normalisedProjects'].add(normalisedProjectName)
        acronyms[normalisedAcronym]['rowNumbers'].add(rowNumber)
        acronyms[normalisedAcronym]['projectsNumbers'].add(projectNumber)

sheetInput.insert_cols(sheetInput.max_column + 1, 4)
sheetInput.cell(row=1, column=sheetInput.max_column - 3, value="Akronim")
sheetInput.cell(row=1, column=sheetInput.max_column - 2, value="Sugerowana nazwa projektu")
sheetInput.cell(row=1, column=sheetInput.max_column - 1, value="Aktualna nazwa projektu")
sheetInput.cell(row=1, column=sheetInput.max_column, value="Wybrany tytuł")

wbOutput = Workbook()
sheetOutput = wbOutput.active
sheetOutput.title = "acronyms"
updatedRows = set()
projectsCounter = 0
acronymsCounter = 0
longDistance = 0
sheetOutput.append(["Source ID", "Acronym", "First Project Name", "Project Name", "Row Number", "Distance from First Project Name"])

for acronym, data in acronyms.items():
    projects = list(data['projects'])
    normalisedProjects = list(data['normalisedProjects'])
    rowNumbers = list(data['rowNumbers'])
    projectsCount = len(projects)
    acronymName = data.get('acronymName', '')
    projectNumbers = list(data['projectsNumbers'])

    if projectsCount > 1:
        firstProject = projects[0]
        firstNormalisedProject = normalisedProjectName[0]
        for i, (project, normalisedProjectName, rowNumber, projectNumber) in enumerate(zip(projects, normalisedProjects, rowNumbers, projectNumbers)):
            projectsCounter += 1
            dist = Levenshtein.distance(firstProject, project)
            sourceId = projectNumber
            sheetOutput.append([sourceId, acronymName, projects[0], project, rowNumber, dist])
            if i > 0 and dist > 49:
                longDistance += 1
                inputRow = sheetInput[rowNumber]
                inputRow[-4].value = acronymName
                inputRow[-3].value = firstProject
                inputRow[-2].value = project
                inputRow[-1].value = ""
                updatedRows.add(int(rowNumber))

wbUpdated = Workbook()
sheetUpdated = wbUpdated.active
header = [cell.value for cell in sheetInput[1]]
sheetUpdated.append(header)
for row in sheetInput.iter_rows(min_row=2, values_only=False):
    rowNumber = row[0].row
    if rowNumber in updatedRows:
        updatedRow = [cell.value for cell in row]
        normalizedAcronym = acronymToNormalized.get(str(updatedRow[-4]).strip(), str(updatedRow[-4]).strip())

        logging.debug(f"Row {rowNumber}: normalizedAcronym = {normalizedAcronym}")
        logging.debug(f"Acronyms dict for this acronym: {acronyms[normalizedAcronym]}")
        logging.debug(f"Updated row: {updatedRow}")

        sanitizedRow = [sanitizeForExcel(cell) for cell in updatedRow]
        sheetUpdated.append(sanitizedRow)

wbOutput.save('data/acronyms.xlsx')
wbUpdated.save('data/input_modified_updated_with_acronyms.xlsx')

print(f"Results have been saved to data/acronyms.xlsx")
print(f"Total projects with multiple acronyms: {projectsCounter}")
print(f"With long distance: {longDistance}")
print(f"Total unique acronyms: {acronymsCounter}")