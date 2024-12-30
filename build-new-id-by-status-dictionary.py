import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import Levenshtein
import logging

logging.basicConfig(level=logging.DEBUG)

needNewIndex = list()

try:
    wbTitles = load_workbook(filename='data/slownik_statusow.xlsx', read_only=False)
    sheetTitlesInput = wbTitles.active
    for rowIndex, row in enumerate(sheetTitlesInput.iter_rows(min_row=2, values_only=True), start=2):
        needNewId = str(row[-3]).strip() if row[-3] is not None else ""
        reportNumber = str(row[0]).strip() if row[-1] is not None else ""

        if needNewId == "TAK":
            needNewIndex.append(reportNumber)
except Exception as e:
    logging.error(f"An error occurred while processing row: {e}")

# Create output workbook
wbNewIdOutput = Workbook()
sheetNewIdOutput = wbNewIdOutput.active
sheetNewIdOutput.title = "Ids"
sheetNewIdOutput.append(["Report number"])
# Write header
for data in needNewIndex:
    reportNumber = data
    sheetNewIdOutput.append([reportNumber])

wbNewIdOutput.save('data/dictionaries/statuses/manual.xlsx')

print(f"Results have been saved to data/dictionaries/statuses/manual.xlsx")