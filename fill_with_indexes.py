import re
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import logging

def sanitizeForExcel(value):
    if isinstance(value, (set, list, tuple)):
        return ', '.join(str(item) for item in value)
    return str(value) if value is not None else ''

def findAcronymInManualDictionaries(projectIndex, acronym):
    returnedValue = acronym
    for index, acronymToUse in manualAcronymsDictionary.items():
        if index == projectIndex:
            returnedValue = acronymToUse
    return returnedValue

def findProjectNameInManualDictionaries(projectIndex, projectName):
    returnedValue = projectName
    for index, projectNameToUse in projectNamesDictionary.items():
        if index == projectIndex:
            returnedValue = projectNameToUse
            break
    return returnedValue

def findProjectNameInDictionaries(projectIndex, projectName):
    returnedValue = projectName
    for index, projectNameToUse in projectNamesDictionary.items():
        if index == projectIndex:
            returnedValue = projectNameToUse
            break
    return returnedValue

def getId(projectIndex, reportIndex):
    global lastId
    searchForIndex = True
    returnedValue = lastId + 1
    for reportId in statusesIndex.items():
        if(reportIndex == reportId):
            searchForIndex = False
            break
    if(searchForIndex):
        for index, id in projectsIds.items():
            if(projectIndex == index):
                returnedValue = id
                break
    if(returnedValue > lastId):
        lastId = lastId + 1
    return returnedValue

logging.basicConfig(level=logging.DEBUG)
projects = defaultdict(lambda: defaultdict(set))
indexes = defaultdict(lambda: defaultdict(set))
projectNamesDictionary = defaultdict(lambda: defaultdict(set))
manualAcronymsDictionary = defaultdict(lambda: defaultdict(set))
manualProjectNamesDictionary = defaultdict(lambda: defaultdict(set))
projects = defaultdict(lambda: defaultdict(set))
statusesIndex = defaultdict(lambda: defaultdict(set))
projectsIds = defaultdict(lambda: defaultdict(set))
lastId = 0

try:
    wbIds = load_workbook(filename='data/dictionaries/index/index.xlsx', read_only=False)
    idsInput = wbIds.active
    for rowIndex, row in enumerate(idsInput.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0])
        projectsIds[index] = int(row[3])
        lastId = int(row[3])
except Exception as e:
    logging.error(e)
    lastId = 0
logging.info("Loaded ids. LastId: " + str(lastId))

try:
    wbDictionaryDistanceProjectNames = load_workbook(filename='data/dictionaries/project_names/based-on-distance.xlsx', read_only=False)
    dictionaryDistanceProjectNamesInput = wbDictionaryDistanceProjectNames.active
    for rowIndex, row in enumerate(dictionaryDistanceProjectNamesInput.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0]) + ';' + str(row[1])
        projectNamesDictionary[index] = str(row[2])
except Exception as e:
    logging.error(e)
    logging.error("Error loading dictionaryDistanceProjectNamesInput")

logging.info("Loaded names")

try:
    wbDictionaryManualProjectNames = load_workbook(filename='data/dictionaries/project_names/manual.xlsx', read_only=False)
    dictionaryManualProjectNamesInput = wbDictionaryManualProjectNames.active
    for rowIndex, row in enumerate(dictionaryManualProjectNamesInput.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0]) + ';' + str(row[1]) + ';' + str(row[2])
        manualProjectNamesDictionary[index] = str(row[3])
except Exception as e:
    logging.error(e)
    logging.error("Error loading dictionaryManualProjectNamesInput")

logging.info("Loaded manual names")

try:
    wbDictionaryManualAcronyms = load_workbook(filename='data/dictionaries/acronyms/manual.xlsx', read_only=False)
    dictionaryManualAcronymsInput = wbDictionaryManualAcronyms.active
    for rowIndex, row in enumerate(dictionaryManualAcronymsInput.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0])
        manualAcronymsDictionary[index] = index
except Exception as e:
    logging.error(e)
    logging.error("Error loading dictionaryManualAcronymsInput")

logging.info("Loaded manual acronyms")

try:
    wbStatuses = load_workbook(filename='data/dictionaries/statuses/manual.xlsx', read_only=False)
    statusesInput = wbStatuses.active
    for rowIndex, row in enumerate(statusesInput.iter_rows(min_row=2, values_only=True), start=2):
        index = str(row[0])
        statusesIndex[index] = str(row[0])
except Exception as e:
    logging.error(e)
    logging.error("Error loading dictionaryStatusesInput")

logging.info("Loaded statuses")

try:
    wbInput = load_workbook(filename='data/input.xlsx', read_only=False)
    logging.info("Loaded")
    sheetInput = wbInput.active
    for rowIndex, row in enumerate(sheetInput.iter_rows(min_row=2, values_only=True), start=2):
        reportNumber = str(row[0]).strip() if row[0] is not None else ""
        normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', str(row[21]).strip().lower()) if row[21] is not None else ""
        normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', str(row[22]).strip().lower()) if row[22] is not None else ""
        projectIndexForManual = normalisedProjectName + ';' + normalisedAcronym + ';' + reportNumber
        projectName = str(row[21]).strip() if row[21] is not None else ""
        acronym = str(row[22]).strip() if row[22] is not None else ""
        projectName = findProjectNameInManualDictionaries(projectIndexForManual, projectName)
        acronym = findAcronymInManualDictionaries(projectIndexForManual, acronym)
        normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', str(projectName).strip().lower())
        normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', str(acronym).strip().lower())
        projectIndex = normalisedProjectName + ';' + normalisedAcronym
        projectName = findProjectNameInDictionaries(projectIndex, projectName)
        normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', str(projectName).strip().lower())
        id = getId(projectIndex, reportNumber)
        inputRow = sheetInput[rowIndex]
        currentId = f"{id:06d}"
        inputRow[20].value = currentId
        if(rowIndex % 100 == 0):
            logging.info("Processed " + str(rowIndex) + " rows and current id: " + currentId)
except Exception as e:
    logging.error(e)
    logging.error("Error loading input")
logging.info("Finished Processed " + str(rowIndex) + " rows")
wbUpdated = Workbook()
sheetUpdated = wbUpdated.active
header = [cell.value for cell in sheetInput[1]]
sheetUpdated.append(header)
for row in sheetInput.iter_rows(min_row=2, values_only=False):
    updatedRow = [cell.value for cell in row]
    sanitizedRow = [sanitizeForExcel(cell) for cell in updatedRow]
    sheetUpdated.append(sanitizedRow)

wbUpdated.save('output/filled.xlsx')
logging.info("Results have been saved to output/filled.xlsx")