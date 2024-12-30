from openpyxl import load_workbook, Workbook
from collections import defaultdict
import Levenshtein
import logging

logging.basicConfig(level=logging.DEBUG)

acronyms = defaultdict(lambda: defaultdict(set))
acronym_to_normalized = {}
def sanitize_for_excel(value):
    if isinstance(value, (set, list, tuple)):
        return ', '.join(str(item) for item in value)
    return str(value) if value is not None else ''

wb = load_workbook(filename='data/input.xlsx', read_only=False)
sheet_input = wb.active
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, values_only=True), start=2):
    projectName = str(row[21]).strip() if row[21] is not None else ""
    normalisedProjectName = re.sub(r'[^a-zA-Z0-9]', '', projectName.lower())
    acronym = str(row[22]).strip() if row[22] is not None else ""
    normalisedAcronym = re.sub(r'[^a-zA-Z0-9]', '', acronym.lower())
    project_index = normalised_project_name + ';' + normalised_acronym
    row_number = str(row_index)
    acronym_to_normalized[acronym] = normalised_acronym
    project_number = str(row[0]).strip() if row[0] is not None else ""
    status = str(row[24]).strip() if row[24] is not None else ""
    institute = str(row[18]).strip() if row[18] is not None else ""

    if project_name and acronym:
        acronyms[project_index]['acronym_name'] = acronym
        acronyms[project_index]['projects'].add(project_name)
        acronyms[project_index]['normalised_projects'].add(normalised_project_name)
        acronyms[project_index]['row_numbers'].add(row_number)
        acronyms[project_index]['projects_numbers'].add(project_number)
        acronyms[project_index]['statuses'].add(status)
        acronyms[project_index]['institutes'].add(institute)

sheet_input.insert_cols(sheet_input.max_column + 1, 6)
sheet_input.cell(row=1, column=sheet_input.max_column - 5, value="Akronim")
sheet_input.cell(row=1, column=sheet_input.max_column - 4, value="Sugerowana nazwa projektu")
sheet_input.cell(row=1, column=sheet_input.max_column - 3, value="Aktualna nazwa projektu")
sheet_input.cell(row=1, column=sheet_input.max_column - 2, value="Wybrany tytuł")
sheet_input.cell(row=1, column=sheet_input.max_column - 1, value="Status")
sheet_input.cell(row=1, column=sheet_input.max_column, value="Instytut")
# Create output workbook
wb_output = Workbook()
sheet_output = wb_output.active
sheet_output.title = "acronyms"
updated_rows = set()
projectsCounter = 0
acronymsCounter = 0
longDistance = 0
sheet_output.append(["Source ID", "Acronym", "First Project Name", "Project Name", "Row Number", "Distance from First Project Name", "Status"])
# Write header
for acronym, data in acronyms.items():
    projects = list(data['projects'])
    normalised_projects = list(data['normalised_projects'])
    row_numbers = list(data['row_numbers'])
    projects_count = len(projects)
    acronym_name = data.get('acronym_name', '')
    project_numbers = list(data['projects_numbers'])
    statuses = list(data['statuses'])
    institutes = list(data['institutes'])

    if projects_count > 1:
        first_project = projects[0]
        first_normalised_project = normalised_project_name[0]
        for i, (project, normalised_project_name, row_number, project_number, status, institute) in enumerate(zip(projects, normalised_projects, row_numbers, project_numbers, statuses, institutes)):
            projectsCounter += 1
            dist = Levenshtein.distance(first_project, project)
            source_id = project_number
            sheet_output.append([source_id, acronym_name, projects[0], project, row_number, dist])
            if status.lower() in ["zakończony", "odrzucony formalnie", "odrzucony merytorycznie", "wycofany", "przerwany", "porzucony"]:
                longDistance += 1
                input_row = sheet_input[row_number]
                input_row[-6].value = acronym_name
                input_row[-5].value = first_project
                input_row[-4].value = project
                input_row[-3].value = ""
                input_row[-2].value = status
                input_row[-1].value = institute
                updated_rows.add(int(row_number))

wb_updated = Workbook()
sheet_updated = wb_updated.active
header = [cell.value for cell in sheet_input[1]]
sheet_updated.append(header)
for row in sheet_input.iter_rows(min_row=2, values_only=False):
    row_number = row[0].row
    if row_number in updated_rows:
        updated_row = [cell.value for cell in row]
        normalized_acronym = acronym_to_normalized.get(str(updated_row[-4]).strip(), str(updated_row[-4]).strip())

        logging.debug(f"Row {row_number}: normalized_acronym = {normalized_acronym}")
        logging.debug(f"Acronyms dict for this acronym: {acronyms[normalized_acronym]}")

        logging.debug(f"Updated row: {updated_row}")

        sanitized_row = [sanitize_for_excel(cell) for cell in updated_row]
        sheet_updated.append(sanitized_row)

wb_output.save('data/acronyms4.xlsx')
wb_updated.save('data/input_e_modified_updated_acronyms_statuses.xlsx')

print(f"Results have been saved to data/acronyms.xlsx")
print(f"Total projects with multiple acronyms: {projectsCounter}")
print(f"With long distance: {longDistance}")
print(f"Total unique acronyms: {acronymsCounter}")